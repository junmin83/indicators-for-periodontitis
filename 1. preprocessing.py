#!/usr/bin/env python
# coding: utf-8

# In[1]:


from __future__ import print_function
import matplotlib.pyplot as plt
import numpy as np
import os
import sys
import pandas as pd
import openpyxl
from sklearn import metrics
import sklearn
import random


# In[2]:


data_location = os.getcwd()+"/health_information_based_features.xlsx"
raw_data = openpyxl.load_workbook(data_location)
raw_data = raw_data.active


# In[3]:


feature_name = ['Sex','Age','Household income','Education','Self-rated stress level','Blood pressure','Glycated hemoglobin',
                'Dental visit within a year','HDL cholesterol','Triglycerides','Abdominal obesity','Fasting blood glucose',
               'Metabolic syndrome','Body mass index','Diabetes','Hs-CRP','Smoking','Use of interdental cleaning aid',

                
label_name = ['NSP','SP']

total_datasize = 13946

features = np.zeros((total_datasize,len(feature_name)))
label = np.zeros((total_datasize,len(label_name)))


# In[4]:


for person_count in range(total_datasize):
    for i in range(4):
        features[person_count,i] = raw_data.cell(row=person_count+2, column=2+i).value
        
    for i in range(4):
        features[person_count,i+4] = raw_data.cell(row=person_count+2, column=7+i).value
        
        
    for i in range(13):
        features[person_count,i+4+4] = raw_data.cell(row=person_count+2, column=13+i).value
        
    if raw_data.cell(row=person_count+2, column=26).value ==1:
        label[person_count,0] = 1  
    elif raw_data.cell(row=person_count+2, column=26).value ==2:
        label[person_count,1] = 1  


# In[5]:


temp_features = features.copy()
temp_label = label.copy()

sequence = np.arange(total_datasize)
random.shuffle(sequence)
for num in range(total_datasize):
    features[num,:] = temp_features[sequence[num],:]
    label[num] = temp_label[sequence[num],:]


# In[43]:


# NOP vs TP
# np.save('dataset_NOPTP\\features.npy',features)
# np.save('dataset_NOPTP\\label.npy',label[:,0]+label[:,1])


# In[9]:


# NSP vs SP
SP_data_num = int(np.sum(label[:,0]+label[:,1]))
print(SP_data_num)
SP_features = np.zeros((SP_data_num,len(feature_name)))
SP_label = np.zeros((SP_data_num,))

temp_count = 0
for person_count in range(total_datasize):
    if not label[person_count,0]+label[person_count,1] ==1:
        continue
    else:
        SP_features[temp_count,:] = features[person_count,:]
        SP_label[temp_count] = label[person_count,1]
        temp_count+=1
    


# In[10]:


# NSP vs SP
# np.save('dataset_NSPSP\\features.npy',SP_features)
# np.save('dataset_NSPSP\\label.npy',SP_label)

